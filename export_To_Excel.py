import openpyxl as xl
import datetime as dt

# Roll number is set according to BIT Mesra standards
def write(rollList):
    print("Data receception by Attendance Tracker Successful!!")
    file_xl="Attendance.xlsx"
    wb = xl.Workbook()
    ws = wb.active
    ws['A1'] = "Roll Number of those present"
    ws['B1'] = dt.datetime.now()
    str = []
    for i in rollList:
        s = i.split("/")
        s1 = s[0].split("_")
        s2 = s[1].split("_")
        str.append([s1[0],s2[0] + " " + s2[1]])
    for i in str:
        ws.append(i)
    wb.save(file_xl)
    print(f"Changes save to: {file_xl}")

# test commands for this function
# write(roll)
